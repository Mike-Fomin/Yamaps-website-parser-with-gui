import json
import os
import os.path
import time
import chromedriver_autoinstaller
import tkinter as tk
import logging
import threading
import traceback
import openpyxl

from datetime import datetime
from tkinter.font import Font
from tkinter import filedialog
from tkinter import INSERT
from selenium import webdriver
from selenium_stealth import stealth
from bs4 import BeautifulSoup
from bs4.element import Tag
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from selenium.common.exceptions import TimeoutException, NoSuchElementException


FLAG: bool = False
FILEPATH: str = ''


class WidgetLogger(logging.Handler):
    def __init__(self, widget):
        logging.Handler.__init__(self)
        self.widget = widget
        self.widget.config(state='disabled')

    def emit(self, record: str) -> None:
        self.widget.config(state='normal')
        # Append message (record) to the widget
        self.widget.insert(INSERT, f"{record}")
        self.widget.see(tk.END)
        self.widget.config(state='disabled')

    def delete(self) -> None:
        self.widget.config(state='normal')
        self.widget.delete(1.0, tk.END)
        self.widget.config(state='disabled')


def openfile() -> None:
    global FILEPATH
    FILEPATH = filedialog.askopenfilename(initialdir=os.path.curdir,
                                          title='Open txt file',
                                          filetypes=(('text files', '*.txt'),)
                                          )
    browse_path.insert(2, FILEPATH)
    button_start.configure(state='normal')
    button_stop.configure(state='normal')


def start_parse() -> None:
    button_start.configure(state='disabled')
    threading.Thread(target=parse).start()


def stop() -> None:
    global FLAG
    FLAG = True
    time.sleep(3)
    button_start.configure(state='normal')


def parse() -> None:
    global FLAG
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")

    # image disable
    options.add_argument('--blink-settings=imagesEnabled=false')

    options.add_argument("--headless")
    options.add_argument("--window-size=1920,1080")

    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    options.page_load_strategy = 'normal'

    driver = webdriver.Chrome(options=options)

    stealth(driver,
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.53 Safari/537.36',
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=False,
            run_on_insecure_origins=False,
            )

    wait = WebDriverWait(driver, 15)

    logtext = WidgetLogger(logger)
    count_log = WidgetLogger(counter_info)

    with open(FILEPATH, 'r', encoding='utf-8') as file:
        websites = list(map(str.rstrip, file.readlines()))
    length: int = len(websites)
    logtext.emit(f"\nОткрытие файла {FILEPATH.split('/')[-1]}. Количество элементов = {length}\n")

    if os.path.exists('cond.json'):
        with open('cond.json', 'r', encoding='utf-8') as json_file:
            json_data = json.load(json_file)
        index: int = json_data['index']
        counter: int = json_data['counter']
        logtext.emit(f"Продолжаем с элемента {json_data['index']} - {json_data['website']}\n")
    else:
        index: int = 0
        counter: int = 0
    count_log.delete()
    count_log.emit(f" {counter}")

    goods: list = []
    try:
        driver.get("https://yandex.ru/maps/moscow/")
        driver.implicitly_wait(7)
        time.sleep(3)

        wait.until(EC.presence_of_element_located((By.TAG_NAME, 'input')))
        input_line = driver.find_element(By.TAG_NAME, 'input')
        for i, website in enumerate(websites, 1):
            try:
                if i < index:
                    continue
                if FLAG:
                    raise KeyboardInterrupt('Приостановка работы!')
                website = website.lower()
                logtext.emit(f"\n{i}/{length} - {website}")
                item: dict = {'index': i, 'website': website, 'counter': counter}

                input_line.click()
                time.sleep(0.2)
                input_line.send_keys(Keys.CONTROL, 'a')
                time.sleep(0.2)
                input_line.send_keys(Keys.DELETE)
                time.sleep(0.2)
                input_line.send_keys(website)
                time.sleep(0.2)
                input_line.send_keys(Keys.ENTER)
                time.sleep(0.2)

                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '[type="submit"]')))
                time.sleep(0.2)

                page = driver.page_source
                soup = BeautifulSoup(page, 'lxml')

                card_selen = driver.find_element(By.CSS_SELECTOR, '.scroll__content')
                card = soup.find(name='div', class_='scroll__content')
                empty = card.find(name='div', class_='search-list-view')
                if empty:
                    continue
                else:
                    title: Tag = card.find(name='h1', attrs={'itemprop': 'name'})
                    categories_block: Tag | None = card.find(name='div', class_='business-categories-view')
                    categories: list = []
                    if categories_block:
                        categories: list[Tag] = categories_block.find_all(name='a')

                    url_block = card.find(name='div', class_='business-urls-view')
                    if url_block:
                        more = url_block.find(name='div', class_='card-feature-view__additional')
                        if more:
                            url_block_selen = card_selen.find_element(By.CSS_SELECTOR, '.business-urls-view')
                            url_block_selen.find_element(By.CSS_SELECTOR, '.card-feature-view__additional').click()
                            time.sleep(0.3)

                            page_add = driver.page_source
                            soup_add = BeautifulSoup(page_add, 'lxml')
                            card_add = soup_add.find(name='div', class_='scroll__content')
                            url_block = card_add.find(name='div', class_='business-urls-view')

                        urls = url_block.find_all(name='a', class_='business-urls-view__link')
                        urls = list(map(lambda x: x.text, urls))
                        if any(map(lambda x: website in x, urls)):
                            logtext.emit(f" ---> Совпадение по email ---> ")
                            reviews_block = card.find(name='div', class_='business-card-title-view__header')
                            reviews = reviews_block.find(name='div', class_='business-header-rating-view__text').text.lower()
                            if reviews == 'ещё нет отзывов':
                                revs = 0
                            else:
                                revs = int(reviews.split()[0].replace('(', '').replace(')', ''))
                            status = card.find(name='div', class_='business-card-title-view__bottom').text.lower()
                            if revs < 2:
                                logtext.emit('не хватает отзывов')
                                continue
                            elif any(map(lambda x: x in status, ['больше не работает', 'временно не работает', 'компания закрыта'])):
                                logtext.emit(f'статус "{status}"')
                                continue
                            else:
                                logtext.emit('добавляю к результату!')

                                title_text: str = title.text.strip()
                                category: str = '; '.join(list(map(lambda x: x.get('title'), categories))) if categories else ''
                                new_item: dict[str, str] = {
                                    'website': website,
                                    'title': title_text,
                                    'category': category
                                }
                                goods.append(new_item)

                                counter += 1
                                count_log.delete()
                                count_log.emit(f" {counter}")

                    else:
                        continue

            except (TimeoutException, NoSuchElementException):
                logtext.emit(f'\nСервер не отвечает. Перезагрузка страницы...')
                driver.refresh()
                driver.implicitly_wait(7)
                time.sleep(3)
                wait.until(EC.presence_of_element_located((By.TAG_NAME, 'input')))
                continue

    except KeyboardInterrupt:
        logtext.emit(f'\nПоследний элемент = {website}, строка {i}')

        with open('cond.json', 'w', encoding='utf-8') as json_file:
            json.dump(item, json_file, indent=4, ensure_ascii=False)

    except Exception:
        print()
        traceback.print_exc()
        logtext.emit(traceback.format_exc())
        logtext.emit(f'\nПоследний элемент = {website}, строка {i}')

        driver.save_screenshot('error.png')
        with open('error.txt', 'a', encoding='utf-8') as error_file:
            error_file.write(f"\n{datetime.now().strftime('%d-%m-%Y %H:%M')}")
            error_file.write(f"\n{traceback.format_exc()}")

        with open('cond.json', 'w', encoding='utf-8') as json_file:
            json.dump(item, json_file, indent=4, ensure_ascii=False)

    else:
        if os.path.exists('cond.json'):
            os.remove('cond.json')
        logtext.emit("\nПарсинг сайтов успешно завершен!")
    finally:
        FLAG = False

        if os.path.exists('Result.xlsx'):
            wb: Workbook = openpyxl.load_workbook('Result.xlsx')
            ws: Worksheet = wb.active
            row: int = 0
            for row, data in enumerate(ws.iter_rows(values_only=True), 1):
                if not data[0]:
                    break
            max_row: int = row + 1

        else:
            wb: Workbook = openpyxl.Workbook()
            ws: Worksheet = wb.active
            max_row: int = 1

        for row, good_item in enumerate(goods, max_row):
            for col, key in enumerate(('website', 'title', 'category'), 1):
                ws.cell(row=row, column=col).value = good_item.get(key, '')
        wb.save('Result.xlsx')

        logtext.emit(f"\nКоличество сохраненных ссылок: {counter}")

        driver.close()
        driver.quit()


if __name__ == '__main__':
    window = tk.Tk()

    window.geometry('700x500')
    window.resizable(0, 0)
    window.title('Website checker')
    window.configure(bg='#333333')

    logger_font = Font(family='Arial', size=10)
    main_font = Font(family='Arial', size=14)
    italic_font = Font(family='Arial', size=12, slant='italic')

    logger = tk.Text(window, bg='grey', state='normal', height=20, width=95, border=5, font=logger_font)
    logger.place(x=10, y=10)
    browse_info = tk.Label(window, text='Укажите путь к файлу:', font=main_font, bg='#333333', fg='white')
    browse_info.place(x=10, y=356)
    browse_path = tk.Entry(window, fg='white', bg='grey', state='normal', width=51, font=italic_font)
    browse_path.place(x=215, y=360)
    button_browse = tk.Button(window, text='Найти', font=main_font, width=22, height=1, bg='light grey',
                              command=openfile)
    button_browse.place(x=400, y=400)
    button_start = tk.Button(window, text='Начать', font=main_font, width=22, height=1, bg='light grey', state='disabled', command=start_parse)
    button_start.place(x=70, y=450)
    button_stop = tk.Button(window, text='Остановить', font=main_font, width=22, height=1, bg='light grey', state='disabled', command=stop)
    button_stop.place(x=400, y=450)
    counter_label = tk.Label(window, text='Счетчик: ', font=main_font, bg='#333333', fg='white')
    counter_label.place(x=10, y=390)
    counter_info = tk.Text(window, bg='grey', state='normal', width=5, height=1, font=main_font)
    counter_info.place(x=100, y=395)

    WidgetLogger(logger).emit('Загрузка chrome драйвера...\n')
    WidgetLogger(logger).emit('Это может занять несколько минут\n')
    chromedriver_autoinstaller.install()
    WidgetLogger(logger).emit('Загрузка chrome драйвера завершена!\n')

    window.mainloop()
